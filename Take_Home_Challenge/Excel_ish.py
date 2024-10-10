from openpyxl import load_workbook
import sys

'''Modify the names into pretty String'''
def getName(name):
    #Find first Dash to split string into First and Last name
    Dash = name.index('-')

    #Get first Name
    firstName = name[:Dash]
    #Uppercase First Letter
    firstName = firstName[0].upper() + firstName[1:]

    #Find index where last name ends
    if '-' in name[Dash+1:]:
        end = Dash + name[Dash+1:].index('-') + 1 
    else:
        end = len(name)-2
        
    lastName = name[Dash+1:end]
    #Uppercase First Letter
    lastName = lastName[0].upper() + lastName[1:]

    #Make full name and return
    newName = firstName + " " + lastName
    return newName

def getEmail(email):
    stop = email.index("\n")
    return(email[:stop])
                       
'''Open Folder and Split information into emails and names'''
def splitInfo(filename):
    count = 0
    sets = 0
    nameDict1 = {}
    nameDict2 = {}
    nameDict3 = {}
    
    with open(filename) as f:
        for x, line in enumerate(f, start=0): #enumerate file into strings
            if line[:6] == '/name/': #Check if string is a name
                name = getName(line[6:])
            elif line[:6] == 'mailto': #Check if string is emal
                email = getEmail(line[7:])
                    
                #If we have the email, we have the name
                if (len(nameDict1) <= 150):
                    nameDict1[name] = email
                elif (len(nameDict2) <= 150):
                    nameDict2[name] = email
                else:
                    nameDict3[name] = email
                sets += 1
                    
            count += 1
    print(count)
    print(sets)
    return(nameDict1, nameDict2, nameDict3)

Dic1, Dic2, Dic3 = splitInfo('finished.txt')

#DOING THE EXCEL SHEET
output_file_name = 'Bruh.xlsx'

wb = load_workbook(output_file_name, data_only=True)
ws = wb['Sheet1']
print("opened workbook")

i = 1
for x,y in Dic1.items():
    ws.cell(i,1).value = (x)
    ws.cell(i,2).value = (y)
    i +=1 
for x,y in Dic2.items():
    ws.cell(i,1).value = (x)
    ws.cell(i,2).value = (y)
    i +=1
for x,y in Dic3.items():
    ws.cell(i,1).value = (x)
    ws.cell(i,2).value = (y)
    i +=1 

ws.cell(3,3).value = ("Balls")

print("edited cells")

wb.save(output_file_name)
wb.close()

print("closed workbook")
